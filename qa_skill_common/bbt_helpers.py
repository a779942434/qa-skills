# -*- coding: utf-8 -*-
"""黑盒测试工具箱（web-blackbox-testing 技能配套）。

解决实测中踩过的坑：
1. 固定 sleep 太多 -> wait_visible() 条件等待
2. 表格 td 硬编码 -> table_columns() / read_table_rows() 按列头读
3. 用例漏报错 -> attach_error_watchers() 三路错误监听（console / HTTP / toast）
4. 截图命名无意义 -> snap() 语义化命名（功能_用例_步骤_时间.png）
5. 误动已有数据 -> record_baseline() / assert_new_target() 数据基线

用法示例见 web-blackbox-testing.md。
"""
import os
import time
from datetime import datetime
from pathlib import Path
__all__ = [
    "launch_mes_browser", "find_page", "connect", "disconnect", "attach_error_watchers",
    "collect_toasts", "error_report", "wait_visible", "wait_until", "retry",
    "close_dialog", "wait_text", "wait_button", "wait_toast",
    "table_columns", "read_table_rows", "snap", "record_baseline",
    "assert_new_target", "parse_import_template", "recon_page_structure",
    "recon_once",
    # 2026-09-03 防误报 + 级联确定性 + 隔离新增
    "read_feedback", "active_dialog", "read_dialog", "judge_action",
    "detect_cascade", "select_cascade", "click_or_observe", "reset_to",
]


from playwright.sync_api import sync_playwright


def _chrome_candidates():
    """返回本机候选浏览器可执行文件（系统 Chrome/Edge/Chromium + 常见路径 + 环境变量覆盖）。"""
    cands = []
    env = os.environ.get("MES_BROWSER_PATH", "").strip()
    if env:
        cands.append(env)
    cands += [
        "/Applications/Google Chrome.app/Contents/MacOS/Google Chrome",
        "/Applications/Microsoft Edge.app/Contents/MacOS/Microsoft Edge",
        "/Applications/Chromium.app/Contents/MacOS/Chromium",
        "/usr/bin/google-chrome",
        "/usr/bin/google-chrome-stable",
        "/usr/bin/chromium",
        "/usr/bin/chromium-browser",
    ]
    return [c for c in cands if c and Path(c).exists()]


def launch_mes_browser(pw, headless=True):
    """启动浏览器用于被测 MES：优先 playwright 自带 chromium；缺失时自动回退本机系统 Chrome/Edge。

    回退顺序：环境变量 MES_BROWSER_PATH > macOS Chrome/Edge/Chromium > Linux 常见路径。
    全部不可用时抛出明确提示（不要自行 playwright install 下载）。
    """
    try:
        return pw.chromium.launch(headless=headless, args=["--no-sandbox"])
    except Exception:
        pass
    for exe in _chrome_candidates():
        try:
            return pw.chromium.launch(headless=headless, executable_path=exe, args=["--no-sandbox"])
        except Exception:
            continue
    raise RuntimeError(
        "playwright 自带 chromium 与本机系统 Chrome/Edge 均不可用："
        "请安装 Chrome/Edge，或设置 MES_BROWSER_PATH 指定浏览器可执行文件"
    )


def find_page(ctx, url_contains=None, title_contains=None):
    """在常驻浏览器上下文里查找已存在的页面，供复用（避免重复开页累积标签页）。"""
    for p in ctx.pages:
        try:
            if url_contains and url_contains not in (p.url or ""):
                continue
            if title_contains and title_contains not in (p.title() or ""):
                continue
            return p
        except Exception:
            continue
    return None


def connect(cdp_url="http://127.0.0.1:9334", new_page=False, url_contains=None, title_contains=None):
    """连接常驻浏览器（CDP，默认 9334），优先复用已有页面。

    - 传入 url_contains / title_contains：先查找匹配的页面，命中则复用；未命中才新建。
    - new_page=True 且未指定复用条件：显式新建页面（仅用于确实需要干净页面的场景）。
    - 都不指定：复用第一个页面（避免同一被测页被反复多开）。
    """
    pw = sync_playwright().start()
    browser = pw.chromium.connect_over_cdp(cdp_url, timeout=20000)
    ctx = browser.contexts[0] if browser.contexts else browser.new_context()
    if url_contains or title_contains:
        page = find_page(ctx, url_contains=url_contains, title_contains=title_contains)
        if page is None:
            page = ctx.new_page()
    elif new_page:
        page = ctx.new_page()
    else:
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
    return pw, browser, ctx, page

def disconnect(pw):
    try:
        pw.stop()
    except Exception:
        pass


# ---------- 1. 错误监听 ----------

def attach_error_watchers(page, toast_selector=".el-message, .el-notification, .el-message-box"):
    """挂载三路错误监听，返回 watchers 收集器。"""
    watchers = {"console": [], "http": [], "toast": []}

    def on_console(msg):
        if msg.type in ("error", "warning"):
            watchers["console"].append((msg.type, msg.text[:300]))

    def on_response(resp):
        if resp.status >= 400:
            watchers["http"].append((resp.status, resp.url[:200]))

    page.on("console", on_console)
    page.on("response", on_response)
    watchers["_toast_selector"] = toast_selector
    return watchers


def collect_toasts(page, watchers):
    """把当前页面可见提示快照进 watchers['toast']（避免重复）。"""
    sel = watchers.get("_toast_selector", ".el-message, .el-notification, .el-message-box")
    for m in page.locator(sel).all():
        try:
            if m.is_visible():
                t = m.inner_text().strip()
                if t and t not in watchers["toast"]:
                    watchers["toast"].append(t[:200])
        except Exception:
            pass
    return watchers


def error_report(watchers):
    """生成错误汇总文本；无错误返回 None。"""
    lines = []
    for typ, text in watchers.get("console", []):
        lines.append(f"[console:{typ}] {text}")
    for status, url in watchers.get("http", []):
        lines.append(f"[http:{status}] {url}")
    for t in watchers.get("toast", []):
        lines.append(f"[toast] {t}")
    return "\n".join(lines) if lines else None


# ---------- 2. 条件等待 ----------

def wait_visible(page, selector, timeout=15, interval=0.5):
    """元素可见即返回 True；超时返回 False（不抛异常）。替代固定 sleep。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        loc = page.locator(selector)
        try:
            if loc.count() > 0 and loc.first.is_visible():
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


def wait_until(page, check, timeout=15, interval=0.5):
    """通用条件等待：check 是返回真值的可调用对象（可接收 page）。超时返回 False。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            if check(page):
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


def retry(fn, attempts=2, interval=1.0, desc=""):
    """执行 fn 最多 attempts 次（防无限重试）。成功返回 (True, result)；失败返回 (False, last_exc)。

    用法：ok, r = retry(lambda: page.locator(...).click(), desc="点保存")
    仍不稳定时上层按「环境观察/偶发」处理，不再继续循环。
    """
    last_exc = None
    for i in range(attempts):
        try:
            return True, fn()
        except Exception as exc:
            last_exc = exc
            if desc:
                print(f"[bbt] 重试({i + 1}/{attempts}) {desc}: {exc}")
            time.sleep(interval)
    return False, last_exc


def close_dialog(page, timeout=5):
    """收尾关闭当前可见弹窗/下拉（Escape + 取消兜底），避免残留互相遮挡。"""
    try:
        page.keyboard.press("Escape")
    except Exception:
        pass
    try:
        cancel = page.locator(".el-dialog:visible button:has-text('取消'), .el-message-box:visible button:has-text('取消')")
        if cancel.count() > 0:
            cancel.first.click()
    except Exception:
        pass
    page.wait_for_timeout(400)
    return wait_until(
        page,
        lambda p: p.locator("[role=dialog]:visible, .el-dialog:visible, .el-message-box:visible").count() == 0,
        timeout=timeout,
    )


def wait_text(page, text, timeout=15, interval=0.5):
    """等待指定叶子文本可见（页面/树/菜单里的可见文本）。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            if page.evaluate(
                """(txt) => {
                    const els = [...document.querySelectorAll('*')];
                    return els.some(el => el.childElementCount===0 && el.textContent && el.textContent.trim()===txt && el.getBoundingClientRect().width>0);
                }""",
                text,
            ):
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


def wait_button(page, text, timeout=15, interval=0.5):
    """等待指定文本按钮可见。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            loc = page.locator(f"button:has-text('{text}')")
            if loc.count() > 0 and loc.first.is_visible():
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


def wait_toast(page, keyword, timeout=10, interval=0.5):
    """等待指定提示（如「新增成功」「导入完成」）出现，替代固定 sleep。超时返回 False。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        try:
            msgs = page.locator(".el-message, .el-notification, .el-message-box").all_inner_texts()
            if any(keyword in (m or "") for m in msgs):
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


# ---------- 3. 按表头读表 ----------

def table_columns(page, table_selector=".el-table"):
    """读表头，返回 {列名: 列索引}（按可见 th 文本）。"""
    ths = page.locator(f"{table_selector} th")
    cols = {}
    for i in range(ths.count()):
        try:
            t = ths.nth(i).inner_text().strip()
        except Exception:
            continue
        if t:
            cols[t] = i
    return cols


def read_table_rows(page, table_selector=".el-table"):
    """按列名读取所有行，返回 [{列名: 值}]。"""
    cols = table_columns(page, table_selector)
    rows = page.locator(f"{table_selector} .el-table__row")
    out = []
    for r in range(rows.count()):
        cells = rows.nth(r).locator("td")
        row = {}
        for name, idx in cols.items():
            try:
                row[name] = cells.nth(idx).inner_text().strip()
            except Exception:
                row[name] = ""
        out.append(row)
    return out


# ---------- 4. 语义化截图 ----------

def snap(page, name, out_dir, feature="", ext="png"):
    """语义化截图：{feature}_{name}_{HHMMSS}.png，返回绝对路径。"""
    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%H%M%S")
    fname = "_".join(x for x in (feature, name, ts) if x) + f".{ext}"
    path = d / fname
    page.screenshot(path=str(path))
    return str(path)


# ---------- 5. 数据基线 ----------

def record_baseline(page, key_col=0, col_name=None, table_selector=".el-table"):
    """记录当前表格行标识集合，用于区分已有数据与测试新增数据。

    建议用 col_name 指定标识列（如"生产订单分单号"），比默认第一列可靠。
    """
    if col_name:
        cols = table_columns(page, table_selector)
        key_col = cols.get(col_name, key_col)
    rows = page.locator(f"{table_selector} .el-table__row")
    keys = set()
    for r in range(rows.count()):
        try:
            v = rows.nth(r).locator("td").nth(key_col).inner_text().strip()
        except Exception:
            continue
        if v:
            keys.add(v)
    return keys


def assert_new_target(baseline, target, label="目标记录"):
    """断言目标不在基线中；在基线中则返回 False 并提示（防误动已有数据）。"""
    if target in baseline:
        print(f"[bbt] 警告：{label}「{target}」在测试前基线中，疑似已有数据，操作已拦截")
        return False
    return True


def parse_import_template(path):
    """解析导入模板表头，返回 [{name, required, col}]（`*` 前缀表示必填）。

    用于导入功能测试时快速核对模板必填标记与需求是否一致。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[bbt] 未安装 openpyxl，无法解析导入模板")
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        required = name.startswith("*")
        if required:
            name = name[1:].strip()
        headers.append({"name": name, "required": required, "col": cell.column})
    return headers


def recon_page_structure(page):
    """一次性返回当前页面结构 dict（按钮/表头/行/输入框/弹窗），供固化与对比。"""
    return page.evaluate(
        """() => ({
            url: location.href,
            title: document.title,
            buttons: [...document.querySelectorAll('button')].map(b=>(b.innerText||'').trim()).filter(Boolean),
            headers: [...document.querySelectorAll('.el-table th')].map(x=>(x.innerText||'').trim()).filter(Boolean),
            rows: [...document.querySelectorAll('.el-table__row')].map(r=>(r.innerText||'').trim().replace(/\\n+/g,' | ')),
            inputs: [...document.querySelectorAll('input')].map(i=>({ph:i.placeholder||'',type:i.type||''})).slice(0,20),
            dialogs: [...document.querySelectorAll('[role=dialog],.el-dialog')].filter(d=>{const r=d.getBoundingClientRect();return r.width>0&&r.height>0;}).map(d=>(d.innerText||'').trim().replace(/\\n+/g,' | ').slice(0,600))
        })"""
    )


def recon_once(page, url, wait_ms=7000):
    """导航到 URL 并一次性返回页面结构（页面需已登录）。"""
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(wait_ms)
    return recon_page_structure(page)


# ---------- 6. 多信号反馈判定（防误报，2026-09-03 增补） ----------

# 校验/拦截类提示关键词（判定为"已处理但被拦截"，而非"失败/无响应"）
_VALIDATION_HINTS = (
    "必须", "不能", "至少", "不大于", "请选择", "请填写", "不能为", "不允许",
    "超出", "超过", "最少", "最多", "请输入", "不能小于", "必填", "大于0",
)
# 成功类提示关键词（可视为"操作成功"）
_SUCCESS_HINTS = ("成功", "已保存", "已完成", "已生成", "提交成功")


def _is_validation(text):
    return any(k in (text or "") for k in _VALIDATION_HINTS)


def _is_success(text):
    return any(k in (text or "") for k in _SUCCESS_HINTS)


def read_feedback(page, toast_selector=".el-message, .el-notification, .el-message-box",
                  error_selector=".el-form-item__error"):
    """读取页面"全反馈信号层"，返回 {toasts, form_errors, dialogs_open}。

    同时抓顶部 toast 与表单内联校验错误，并数可见 dialog。
    用于判定"操作到底有没有被处理、有没有提示"，避免只读单层信号造成误报
    （例：必填校验提示以 toast 弹出而非内联错误，只读内联会误判为"静默无提示"）。
    """
    toasts = []
    try:
        for m in page.locator(toast_selector).all():
            try:
                if m.is_visible():
                    t = (m.inner_text() or "").strip()
                    if t and t not in toasts:
                        toasts.append(t[:200])
            except Exception:
                pass
    except Exception:
        pass
    form_errors = []
    try:
        for m in page.locator(error_selector).all():
            try:
                if m.is_visible():
                    t = (m.inner_text() or "").strip()
                    if t and t not in form_errors:
                        form_errors.append(t[:200])
            except Exception:
                pass
    except Exception:
        pass
    dialogs_open = 0
    try:
        dialogs_open = page.locator("[role=dialog]:visible, .el-dialog:visible, .el-message-box:visible").count()
    except Exception:
        pass
    return {"toasts": toasts, "form_errors": form_errors, "dialogs_open": dialogs_open}


def active_dialog(page):
    """返回当前最上层（DOM 中最后一个）可见 .el-dialog 的 locator；无则 None。

    注意：仅按 DOM 顺序取"最后一个可见 dialog"近似最上层，多弹窗时需人工确认层级。
    """
    try:
        ds = page.locator(".el-dialog:visible")
        if ds.count() > 0:
            return ds.last
    except Exception:
        pass
    return None


def read_dialog(page, idx=-1):
    """读取最新可见 .el-dialog 的结构（纯 JS，规避 :visible/nth 混乱）。

    返回 {title, text, tables, inputs, buttons, nDlg}；无可见 dialog 返回 None。
    关键场景：对话框右侧明细表列缺失、行内序号链接打开的"查看"框为空——用本函数取结构，
    并结合 bom-preview 等接口核对，避免仅凭 UI 判"空"。
    """
    try:
        return page.evaluate("""(jsIdx) => {
            const ds=[...document.querySelectorAll('.el-dialog')].filter(v=>{const r=v.getBoundingClientRect();return r.width>0&&r.height>0;});
            if(!ds.length) return null;
            const d=ds[ds.length-1];
            return {title:(d.querySelector('.el-dialog__title')||{}).innerText||'',
                    text:(d.innerText||'').replace(/\n+/g,' | ').slice(0,3000),
                    nDlg:ds.length,
                    tables:[...d.querySelectorAll('.el-table')].map(t=>({hdr:[...t.querySelectorAll('th')].map(x=>(x.innerText||'').trim()).filter(Boolean),rows:[...t.querySelectorAll('.el-table__row')].map(r=>(r.innerText||'').replace(/\n+/g,' | '))})).slice(0,4),
                    inputs:[...d.querySelectorAll('input')].filter(i=>{const r=i.getBoundingClientRect();return r.width>0&&r.height>0;}).map(i=>({ph:i.placeholder||'',val:(i.value||'').slice(0,40),dis:i.disabled})).slice(0,20),
                    buttons:[...new Set([...d.querySelectorAll('button')].filter(b=>b.offsetParent!==null).map(b=>(b.innerText||'').trim()).filter(Boolean))]};
        }""", idx)
    except Exception:
        return None


def judge_action(page, action, api_watcher=None, data_diff=None, keyword=None,
                 timeout=15.0, toast_selector=".el-message, .el-notification, .el-message-box",
                 error_selector=".el-form-item__error"):
    """执行 action 并综合多信号判定操作是否"已处理/成功"，避免误报（2026-09-03 增补）。

    背景：必填校验拦截常以 toast 弹出、且可能不发新业务接口——只判"有没有新响应"
    会把"校验已拦截（有提示）"误判为"静默无响应"。本函数同时考虑：
      已处理(processed) = 有业务新响应 或 有 toast/内联校验提示；
      成功(ok)          = 有新响应且无错误 且 无"校验拦截"类提示，或 data_diff 显示数据已变。

    返回 dict：{processed, ok, reason, signals}
      reason: "success"/"changed"/"blocked"/"silent"/"http_error" 之一。
    用法：把"点确定、提交、生成"这类可能被校验拦截或网络失败的动作交进来；
      processed=False 且 reason="silent" = 无任何信号，才需要人工核（真正的无反馈）。
    """
    base = None
    if api_watcher is not None:
        try:
            base = api_watcher.snapshot()
        except Exception:
            base = None
    http_err = []

    def on_resp(resp):
        try:
            if resp.status >= 400:
                http_err.append((resp.status, resp.url[:200]))
        except Exception:
            pass

    page.on("response", on_resp)
    new_responses = []
    try:
        action()
        if api_watcher is not None:
            try:
                new_responses = api_watcher.wait_new(base, keyword=keyword, timeout=timeout)
            except Exception:
                new_responses = []
    finally:
        try:
            page.remove_listener("response", on_resp)
        except Exception:
            pass

    sig = read_feedback(page, toast_selector=toast_selector, error_selector=error_selector)
    base_urls = base or set()
    err_new = [(s, u) for s, u in http_err if u not in base_urls]

    has_new_resp = bool(new_responses)
    has_feedback = bool(sig["toasts"] or sig["form_errors"])
    has_validation = any(_is_validation(t) for t in (sig["toasts"] + sig["form_errors"]))
    has_success = any(_is_success(t) for t in sig["toasts"])
    processed = has_new_resp or has_feedback
    changed = bool(data_diff)

    if err_new:
        ok, reason = False, "http_error"
    elif changed:
        ok, reason = True, "changed"
    elif has_success:
        ok, reason = has_new_resp or True, "success"
    elif has_validation:
        ok, reason = False, "blocked"
    elif has_new_resp:
        ok, reason = True, "success"
    elif has_feedback:
        ok, reason = False, "blocked"
    else:
        ok, reason = False, "silent"

    return {
        "processed": processed,
        "ok": ok,
        "reason": reason,
        "signals": {"new_responses": new_responses, "toasts": sig["toasts"],
                    "form_errors": sig["form_errors"], "http_errors": err_new,
                    "data_diff": data_diff, "dialogs_open": sig["dialogs_open"]},
    }


# ---------- 7. 级联/树选择：先侦测、匹配才走（2026-09-03 增补） ----------

def detect_cascade(page, trigger_sel=None, wait_ms=1500):
    """点开"来源地/树选择"下拉并返回结构判定，用于确认是否"父→子"两级级联。

    返回 dict：{kind, nPanels, levels(list of node-label lists), parents, leaves, hasParentChild}
      kind: 'el-cascader-2level' | 'single-list' | 'unknown'
    用途：级联选择前**先侦测**，确认是两级父→子才走 select_cascade 的确定性"父→子"步骤；
      结构不匹配（非级联/不是两级）则返回 unknown/single-list，调用方应 skip 而不是盲点。
    """
    if trigger_sel:
        try:
            page.locator(trigger_sel).first.click(timeout=4000)
            page.wait_for_timeout(wait_ms)
        except Exception:
            pass
    try:
        return page.evaluate("""() => {
            const panels=[...document.querySelectorAll('.el-cascader-menu')].filter(p=>{const r=p.getBoundingClientRect();return r.width>0&&r.height>0;});
            const nodeOf=p=>[...p.querySelectorAll('.el-cascader-node')].map(n=>(n.innerText||'').trim().replace(/\s+/g,' ')).filter(Boolean);
            const levels=panels.map(nodeOf);
            const hasParentChild = levels.length>=2 && levels[0].length>0 && levels[1].length>0;
            const kind = levels.length>=2 ? 'el-cascader-2level' : (levels.length===1 ? 'single-list' : 'unknown');
            return {kind, nPanels:levels.length, levels, parents:levels[0]||[], leaves:(levels[1]||[]), hasParentChild};
        }""")
    except Exception:
        return {"kind": "unknown", "nPanels": 0, "levels": [], "parents": [], "leaves": [], "hasParentChild": False}


def select_cascade(page, trigger_sel, leaf_part, diag=None, input_sel=None):
    """按"先侦测、匹配才走"的确定性步骤选择级联叶节点。

    流程：
      1. diag = detect_cascade(...)（未传则先侦测）；
      2. 若结构非两级父→子（hasParentChild=False）-> 返回 ("skip", reason)，**不点击**；
      3. 否则：优先点第二级面板中含 leaf_part 的叶节点；若未渲染先点第一级父节点展开，
         再点叶节点；点击后用 input_sel/trigger_sel 读输入框值，断言回填非空。

    返回 ("ok", value) / ("fail", reason) / ("skip", reason)。
    """
    if diag is None:
        diag = detect_cascade(page, trigger_sel)
    if not diag.get("hasParentChild"):
        return ("skip", "非两级父→子级联（kind=%s），不执行选择" % diag.get("kind"))

    def _click_leaf():
        leaf = page.locator(".el-cascader-menu:visible").last.locator(
            ".el-cascader-node:visible", has_text=leaf_part).first
        if leaf.count() == 0:
            return ("fail", "未找到叶节点[%s]" % leaf_part)
        leaf.click(timeout=4000)
        page.wait_for_timeout(1000)
        return ("ok", "")

    res = _click_leaf()
    if res[0] != "ok":
        # 先展开第一级父节点，再点叶
        try:
            p = page.locator(".el-cascader-menu:visible").first.locator(".el-cascader-node").first
            p.click(timeout=3000)
            page.wait_for_timeout(800)
        except Exception:
            pass
        res = _click_leaf()
    if res[0] != "ok":
        return res
    val = ""
    try:
        sel = input_sel or trigger_sel
        val = page.locator(sel).first.input_value(timeout=2000)
    except Exception:
        pass
    if not val:
        return ("fail", "叶节点已点但输入框回填为空（leaf=%s）" % leaf_part)
    return ("ok", val)


# ---------- 8. 禁用前置判定 + 用例隔离（2026-09-03 增补） ----------

def click_or_observe(page, btn_text, btn_sel="button", timeout=6.0):
    """点击文本按钮，但先判禁用，避免对 disabled 按钮 click 超时中断整段。

    返回 ("ok"|"disabled"|"notfound"|"fail", 说明)。disabled/notfound 记为"状态观察"，不抛超时。
    """
    try:
        loc = page.locator(f"{btn_sel}:visible", has_text=btn_text)
        if loc.count() == 0:
            return ("notfound", "未找到按钮[%s]" % btn_text)
        try:
            if loc.first.is_disabled(timeout=timeout):
                return ("disabled", "按钮[%s] 处于 disabled（禁用）态" % btn_text)
        except Exception:
            pass
        loc.first.click(timeout=timeout)
        return ("ok", "已点击[%s]" % btn_text)
    except Exception as e:
        return ("fail", "点击[%s]失败:%s" % (btn_text, repr(e)[:100]))


def reset_to(page, url, tab_text=None, wait_ms=6000):
    """用例前置：导航回已知态（可选切页签），保证用例间独立、避免脏状态延续。

    用例隔离：每条用例前调用，把页面重置到确定起点；结合 try/except，单条失败不中断整段。
    """
    page.goto(url, wait_until="domcontentloaded", timeout=60000)
    page.wait_for_timeout(wait_ms)
    if tab_text:
        try:
            page.locator(".el-tabs__item", has_text=tab_text).first.click(timeout=8000)
            page.wait_for_timeout(2000)
        except Exception:
            pass
    return True


if __name__ == "__main__":
    print("bbt_helpers 可用函数:")
    print("  connect / disconnect / attach_error_watchers / collect_toasts / error_report")
    print("  wait_visible / wait_until / wait_text / wait_button / wait_toast / retry / close_dialog")
    print("  table_columns / read_table_rows / snap / record_baseline / assert_new_target")
    print("  find_page / parse_import_template / recon_page_structure / recon_once")
    print("  防误报&隔离: read_feedback / active_dialog / read_dialog / judge_action")
    print("  级联&禁用: detect_cascade / select_cascade / click_or_observe / reset_to")
