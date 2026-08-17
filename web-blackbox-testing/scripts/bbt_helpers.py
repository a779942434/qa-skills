# -*- coding: utf-8 -*-
"""黑盒测试工具箱（web-blackbox-testing 技能配套）。

解决实测中踩过的坑：
1. 固定 sleep 太多 -> wait_visible() 条件等待
2. 表格 td 硬编码 -> table_columns() / read_table_rows() 按列头读
3. 用例漏报错 -> attach_error_watchers() 三路错误监听（console / HTTP / toast）
4. 截图命名无意义 -> snap() 语义化命名（功能_用例_步骤_时间.png）
5. 误动已有数据 -> record_baseline() / assert_new_target() 数据基线

用法示例见 web-blackbox-testing.md。
"""
import time
from datetime import datetime
from pathlib import Path

from playwright.sync_api import sync_playwright


def find_page(ctx, url_contains=None, title_contains=None):
    """在常驻浏览器上下文里查找已存在的页面，供复用（避免重复开页累积标签页）。"""
    for p in ctx.pages:
        try:
            if url_contains and url_contains not in (p.url or ""):
                continue
            if title_contains and title_contains not in (p.title() or ""):
                continue
            return p
        except Exception:
            continue
    return None


def connect(cdp_url="http://127.0.0.1:9334", new_page=False, url_contains=None, title_contains=None):
    """连接常驻浏览器（CDP，默认 9334），优先复用已有页面。

    - 传入 url_contains / title_contains：先查找匹配的页面，命中则复用；未命中才新建。
    - new_page=True 且未指定复用条件：显式新建页面（仅用于确实需要干净页面的场景）。
    - 都不指定：复用第一个页面（避免同一被测页被反复多开）。
    """
    pw = sync_playwright().start()
    browser = pw.chromium.connect_over_cdp(cdp_url, timeout=20000)
    ctx = browser.contexts[0] if browser.contexts else browser.new_context()
    if url_contains or title_contains:
        page = find_page(ctx, url_contains=url_contains, title_contains=title_contains)
        if page is None:
            page = ctx.new_page()
    elif new_page:
        page = ctx.new_page()
    else:
        page = ctx.pages[0] if ctx.pages else ctx.new_page()
    return pw, browser, ctx, page

def disconnect(pw):
    try:
        pw.stop()
    except Exception:
        pass


# ---------- 1. 错误监听 ----------

def attach_error_watchers(page, toast_selector=".el-message, .el-notification, .el-message-box"):
    """挂载三路错误监听，返回 watchers 收集器。"""
    watchers = {"console": [], "http": [], "toast": []}

    def on_console(msg):
        if msg.type in ("error", "warning"):
            watchers["console"].append((msg.type, msg.text[:300]))

    def on_response(resp):
        if resp.status >= 400:
            watchers["http"].append((resp.status, resp.url[:200]))

    page.on("console", on_console)
    page.on("response", on_response)
    watchers["_toast_selector"] = toast_selector
    return watchers


def collect_toasts(page, watchers):
    """把当前页面可见提示快照进 watchers['toast']（避免重复）。"""
    sel = watchers.get("_toast_selector", ".el-message, .el-notification, .el-message-box")
    for m in page.locator(sel).all():
        try:
            if m.is_visible():
                t = m.inner_text().strip()
                if t and t not in watchers["toast"]:
                    watchers["toast"].append(t[:200])
        except Exception:
            pass
    return watchers


def error_report(watchers):
    """生成错误汇总文本；无错误返回 None。"""
    lines = []
    for typ, text in watchers.get("console", []):
        lines.append(f"[console:{typ}] {text}")
    for status, url in watchers.get("http", []):
        lines.append(f"[http:{status}] {url}")
    for t in watchers.get("toast", []):
        lines.append(f"[toast] {t}")
    return "\n".join(lines) if lines else None


# ---------- 2. 条件等待 ----------

def wait_visible(page, selector, timeout=15, interval=0.5):
    """元素可见即返回 True；超时返回 False（不抛异常）。替代固定 sleep。"""
    deadline = time.time() + timeout
    while time.time() < deadline:
        loc = page.locator(selector)
        try:
            if loc.count() > 0 and loc.first.is_visible():
                return True
        except Exception:
            pass
        time.sleep(interval)
    return False


# ---------- 3. 按表头读表 ----------

def table_columns(page, table_selector=".el-table"):
    """读表头，返回 {列名: 列索引}（按可见 th 文本）。"""
    ths = page.locator(f"{table_selector} th")
    cols = {}
    for i in range(ths.count()):
        try:
            t = ths.nth(i).inner_text().strip()
        except Exception:
            continue
        if t:
            cols[t] = i
    return cols


def read_table_rows(page, table_selector=".el-table"):
    """按列名读取所有行，返回 [{列名: 值}]。"""
    cols = table_columns(page, table_selector)
    rows = page.locator(f"{table_selector} .el-table__row")
    out = []
    for r in range(rows.count()):
        cells = rows.nth(r).locator("td")
        row = {}
        for name, idx in cols.items():
            try:
                row[name] = cells.nth(idx).inner_text().strip()
            except Exception:
                row[name] = ""
        out.append(row)
    return out


# ---------- 4. 语义化截图 ----------

def snap(page, name, out_dir, feature="", ext="png"):
    """语义化截图：{feature}_{name}_{HHMMSS}.png，返回绝对路径。"""
    d = Path(out_dir)
    d.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("%H%M%S")
    fname = "_".join(x for x in (feature, name, ts) if x) + f".{ext}"
    path = d / fname
    page.screenshot(path=str(path))
    return str(path)


# ---------- 5. 数据基线 ----------

def record_baseline(page, key_col=0, col_name=None, table_selector=".el-table"):
    """记录当前表格行标识集合，用于区分已有数据与测试新增数据。

    建议用 col_name 指定标识列（如"生产订单分单号"），比默认第一列可靠。
    """
    if col_name:
        cols = table_columns(page, table_selector)
        key_col = cols.get(col_name, key_col)
    rows = page.locator(f"{table_selector} .el-table__row")
    keys = set()
    for r in range(rows.count()):
        try:
            v = rows.nth(r).locator("td").nth(key_col).inner_text().strip()
        except Exception:
            continue
        if v:
            keys.add(v)
    return keys


def assert_new_target(baseline, target, label="目标记录"):
    """断言目标不在基线中；在基线中则返回 False 并提示（防误动已有数据）。"""
    if target in baseline:
        print(f"[bbt] 警告：{label}「{target}」在测试前基线中，疑似已有数据，操作已拦截")
        return False
    return True


def parse_import_template(path):
    """解析导入模板表头，返回 [{name, required, col}]（`*` 前缀表示必填）。

    用于导入功能测试时快速核对模板必填标记与需求是否一致。
    """
    try:
        from openpyxl import load_workbook
    except ImportError:
        print("[bbt] 未安装 openpyxl，无法解析导入模板")
        return []
    wb = load_workbook(path, data_only=True)
    ws = wb.active
    headers = []
    for cell in ws[1]:
        if cell.value is None:
            continue
        name = str(cell.value).strip()
        required = name.startswith("*")
        if required:
            name = name[1:].strip()
        headers.append({"name": name, "required": required, "col": cell.column})
    return headers


if __name__ == "__main__":
    print("bbt_helpers 可用函数:")
    print("  connect / disconnect / attach_error_watchers / collect_toasts / error_report")
    print("  wait_visible / table_columns / read_table_rows / snap / record_baseline / assert_new_target")
    print("  find_page / parse_import_template")
